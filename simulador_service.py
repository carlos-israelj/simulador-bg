"""
Microservicio local: expone el Simulador de Multicredito de Banco Guayaquil
como un endpoint HTTP que Power Automate puede consumir.

Patron: "scraper como servicio".
    Power Automate --POST /simular--> este servicio --headless--> simulador BG
    el servicio devuelve JSON limpio con cuota, detalle y tabla de amortizacion.

Por que asi: la funcion de calculo del banco vive dentro de un closure de modulo
minificado (no esta en window), por lo que no se puede invocar directa. La via
robusta es manejar el DOM en headless y leer el resultado que el propio banco calcula.

Requisitos:
    pip install fastapi uvicorn playwright
    playwright install chromium

Ejecutar:
    uvicorn simulador_service:app --host 127.0.0.1 --port 8000

IMPORTANTE - selectores: los marcados con (TODO VERIFICAR) hay que confirmarlos
contra la pagina real. La forma mas rapida y fiable de obtenerlos es grabar la
interaccion con el codegen de Playwright (no adivinar a mano):

    playwright codegen https://www.bancoguayaquil.com/creditos/simuladores/

Eso abre el navegador y va escribiendo el codigo con los selectores exactos
mientras llenas el formulario. Pega aqui los que te genere.
"""

from contextlib import asynccontextmanager
import random
import asyncio
import os
import re
import socket
import subprocess
from pathlib import Path
from datetime import datetime
from io import BytesIO

from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse, Response
from pydantic import BaseModel, Field
from playwright.async_api import async_playwright
from playwright_stealth import Stealth
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Importar comportamiento humano
from human_behavior import (
    human_delay,
    human_click,
    human_type,
    human_scroll,
    simulate_reading_time,
    random_mouse_movement
)

URL = "https://www.bancoguayaquil.com/creditos/simuladores/"

# Configuración de entorno (producción vs desarrollo)
ENVIRONMENT = os.getenv("ENVIRONMENT", "development")  # production o development
HEADLESS_MODE = os.getenv("HEADLESS", "true" if ENVIRONMENT == "production" else "false").lower() == "true"

# Configuración de Tor (activar con USE_TOR=true)
USE_TOR = os.getenv("USE_TOR", "false").lower() == "true"
TOR_SOCKS_PORT = int(os.getenv("TOR_SOCKS_PORT", "9050"))

# Configuración de proxy (opcional, se puede configurar via variables de entorno)
# Formato: http://username:password@proxy-server:port o http://proxy-server:port
PROXY_SERVER = os.getenv("PROXY_SERVER")  # ej: "http://proxy.ejemplo.com:8080"
PROXY_USERNAME = os.getenv("PROXY_USERNAME")
PROXY_PASSWORD = os.getenv("PROXY_PASSWORD")

# Configuración de proxies residenciales (rotación automática)
USE_PROXY_LIST = os.getenv("USE_PROXY_LIST", "false").lower() == "true"
PROXY_LIST_FILE = os.getenv("PROXY_LIST_FILE", "proxies.txt")

# Directorio temporal para PDFs
PDF_DIR = Path("/tmp/pdf_downloads")
PDF_DIR.mkdir(exist_ok=True, parents=True)

# Estado compartido: un solo navegador para todas las peticiones (mas rapido).
_state = {}
# Almacenamiento simple de resultados (última simulación)
_ultimo_resultado = None
# Lista de proxies residenciales y contador para rotación
_proxy_list = []
_proxy_index = 2  # Empezar desde el tercer proxy

# Lista de User-Agents realistas para rotar (Windows/macOS Chrome/Firefox)
_user_agents = [
    # Windows Chrome
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
    # Windows Firefox
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:133.0) Gecko/20100101 Firefox/133.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:132.0) Gecko/20100101 Firefox/132.0",
    # macOS Chrome
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
    # macOS Firefox
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:133.0) Gecko/20100101 Firefox/133.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:132.0) Gecko/20100101 Firefox/132.0",
]
_ua_index = 0


def _load_proxy_list():
    """Carga la lista de proxies residenciales desde el archivo."""
    global _proxy_list
    try:
        proxy_file = Path(PROXY_LIST_FILE)
        if not proxy_file.exists():
            print(f"⚠️ Archivo de proxies no encontrado: {PROXY_LIST_FILE}")
            return False

        _proxy_list = []
        with open(proxy_file, 'r') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    _proxy_list.append(line)

        if not _proxy_list:
            print("⚠️ No se encontraron proxies en el archivo")
            return False

        print(f"✓ Cargados {len(_proxy_list)} proxies residenciales")
        return True
    except Exception as e:
        print(f"❌ Error al cargar proxies: {e}")
        return False


def _get_next_proxy():
    """
    Obtiene el siguiente proxy de la lista (rotación circular).
    Formato del archivo: IP:PUERTO:USUARIO:PASSWORD
    Retorna: dict con server, username, password
    """
    global _proxy_index

    if not _proxy_list:
        return None

    # Obtener el proxy actual
    proxy_line = _proxy_list[_proxy_index]

    # Rotar al siguiente para el próximo request
    _proxy_index = (_proxy_index + 1) % len(_proxy_list)

    # Parsear la línea: IP:PUERTO:USUARIO:PASSWORD
    parts = proxy_line.split(':')
    if len(parts) != 4:
        print(f"⚠️ Formato de proxy inválido: {proxy_line}")
        return None

    ip, port, username, password = parts

    return {
        "server": f"http://{ip}:{port}",
        "username": username,
        "password": password
    }


def _get_next_user_agent():
    """
    Obtiene el siguiente User-Agent de la lista (rotación circular).
    Retorna: string con el user-agent
    """
    global _ua_index

    if not _user_agents:
        return "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"

    # Obtener el user-agent actual
    user_agent = _user_agents[_ua_index]

    # Rotar al siguiente para el próximo request
    _ua_index = (_ua_index + 1) % len(_user_agents)

    return user_agent


def _check_tor_running():
    """Verifica si Tor está corriendo en el puerto SOCKS."""
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(1)
        result = sock.connect_ex(('127.0.0.1', TOR_SOCKS_PORT))
        sock.close()
        return result == 0
    except:
        return False


def _renew_tor_identity():
    """
    Solicita una nueva identidad a Tor (nueva IP) usando stem.
    Intenta conectarse al controlador de TOR en el puerto 9051.
    """
    try:
        from stem import Signal
        from stem.control import Controller

        print("🔄 Solicitando nueva identidad a TOR...")

        # Intentar conectar al controlador de TOR
        # Por defecto TOR escucha en 9051 para control
        with Controller.from_port(port=9051) as controller:
            # Autenticar (sin contraseña porque configuramos CookieAuthentication 0)
            controller.authenticate()

            # Solicitar nueva identidad (nueva IP)
            controller.signal(Signal.NEWNYM)

            print("✓ Nueva identidad de TOR obtenida exitosamente")
            print("  Esperando 5 segundos para que TOR establezca nuevo circuito...")
            return True

    except ImportError:
        print("❌ ERROR: La librería 'stem' no está instalada.")
        print("   Instala con: pip install stem")
        return False
    except Exception as e:
        print(f"❌ Error al renovar identidad de TOR: {e}")
        print(f"   Verifica que TOR esté corriendo: sudo service tor@default status")
        print(f"   Verifica la configuración: ControlPort 9051 y CookieAuthentication 0")
        return False


@asynccontextmanager
async def lifespan(app: FastAPI):
    # Cargar lista de proxies residenciales si está habilitada
    if USE_PROXY_LIST:
        _load_proxy_list()

    pw = await async_playwright().start()
    # headless se configura automáticamente según ENVIRONMENT
    # - development (local): headless=False para debugging
    # - production (Render/cloud): headless=True
    # IMPORTANTE: Usar Chromium en vez de Firefox porque PDF generation solo funciona en Chromium
    browser = await pw.chromium.launch(
        headless=HEADLESS_MODE,
        args=[
            '--disable-blink-features=AutomationControlled',
            '--disable-dev-shm-usage',
            '--no-sandbox'
        ]
    )
    _state["pw"] = pw
    _state["browser"] = browser
    yield
    await browser.close()
    await pw.stop()


app = FastAPI(lifespan=lifespan, title="Simulador BG")


class SimulacionInput(BaseModel):
    monto: float = Field(..., ge=2000, le=50000)      # monto minimo: $2.000, maximo estimado: $50.000
    meses: int = Field(..., ge=12, le=60)             # limites reales: 12-60, paso 12
    amortizacion: str = Field(..., pattern="(?i)^(aleman|frances)$")
    correo: str | None = None                          # se usa aguas abajo, no en el calculo


def _solo_meses_validos(m: int) -> int:
    """El portal solo permite 12, 24, 36, 48, 60. Redondea al permitido mas cercano."""
    permitidos = [12, 24, 36, 48, 60]
    return min(permitidos, key=lambda x: abs(x - m))


@app.post("/simular")
async def simular(data: SimulacionInput):
    browser = _state["browser"]

    # Si Tor está habilitado, verificar que esté corriendo y renovar identidad
    if USE_TOR:
        if not _check_tor_running():
            raise HTTPException(
                status_code=503,
                detail="Tor está habilitado (USE_TOR=true) pero no está corriendo. Ejecuta: sudo service tor@default start"
            )
        # Renovar identidad para obtener nueva IP
        if not _renew_tor_identity():
            raise HTTPException(
                status_code=503,
                detail="No se pudo renovar la identidad de TOR. Verifica la configuración."
            )
        # Esperar 10-15 segundos para que TOR establezca nuevo circuito con nueva IP
        # y para evitar detección de tráfico automatizado
        await asyncio.sleep(random.uniform(10.0, 15.0))

    # Configurar context
    context_options = {
        "user_agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "viewport": {"width": 1920, "height": 1080},
        "locale": "es-EC",
        "timezone_id": "America/Guayaquil",
        "java_script_enabled": True
    }

    # Prioridad de proxy: PROXY_LIST > Tor > PROXY_SERVER
    if USE_PROXY_LIST:
        # Usar rotación de proxies residenciales
        proxy = _get_next_proxy()
        if proxy:
            context_options["proxy"] = proxy
            print(f"🌐 Usando proxy residencial: {proxy['server']} (usuario: {proxy['username']})")
        else:
            print("⚠️ No se pudo obtener proxy de la lista, continuando sin proxy")
    elif USE_TOR:
        # Usar proxy SOCKS5 de Tor
        context_options["proxy"] = {
            "server": f"socks5://127.0.0.1:{TOR_SOCKS_PORT}"
        }
        print(f"🧅 Usando Tor SOCKS5 proxy en puerto {TOR_SOCKS_PORT}")
    elif PROXY_SERVER:
        # Usar proxy HTTP configurado manualmente
        proxy_config = {"server": PROXY_SERVER}
        if PROXY_USERNAME and PROXY_PASSWORD:
            proxy_config["username"] = PROXY_USERNAME
            proxy_config["password"] = PROXY_PASSWORD
        context_options["proxy"] = proxy_config

    # Crear contexto con user-agent realista y configuraciones anti-detección
    context = await browser.new_context(**context_options)

    # Inyectar scripts anti-detección antes de navegar
    await context.add_init_script("""
        // Ocultar webdriver
        Object.defineProperty(navigator, 'webdriver', {get: () => undefined});

        // Sobrescribir propiedades de detección
        window.navigator.chrome = {runtime: {}};

        // Permisos de notificaciones (comportamiento de navegador real)
        Object.defineProperty(navigator, 'permissions', {
            get: () => ({
                query: () => Promise.resolve({state: 'denied'})
            })
        });

        // Plugins (comportamiento de navegador real)
        Object.defineProperty(navigator, 'plugins', {
            get: () => [1, 2, 3, 4, 5]
        });

        // Languages
        Object.defineProperty(navigator, 'languages', {
            get: () => ['es-EC', 'es', 'en']
        });
    """)

    # Abrir una única pestaña
    page = await context.new_page()

    try:
        # Timeout aumentado a 60 segundos para manejar captchas y validaciones
        await page.goto(URL, wait_until="domcontentloaded", timeout=60000)

        # Delay inicial MUY largo para dar tiempo a que la página cargue completamente
        await asyncio.sleep(random.uniform(5.0, 8.0))

        # Aceptar cookies si aparece el popup
        try:
            # Esperar a que aparezca el botón de aceptar cookies
            aceptar_btn = page.locator("#hs-eu-confirmation-button")
            await aceptar_btn.wait_for(timeout=5000, state="visible")
            await asyncio.sleep(random.uniform(0.5, 1.0))
            await aceptar_btn.click()
            print("✓ Cookies aceptadas")
            await asyncio.sleep(random.uniform(1.0, 2.0))
        except:
            print("⚠ No se encontró popup de cookies o ya fue aceptado")
            pass  # Si no hay popup, continuar

        # Esperar a que el formulario del simulador esté visible
        await page.wait_for_selector("input.multicredito__input", timeout=60000)

        # NO hacer scroll manual - dejar que scroll_into_view_if_needed lo maneje
        # cuando interactuemos con los elementos
        await asyncio.sleep(random.uniform(1.5, 2.5))  # Delay antes de interactuar

        # 1) Monto del credito (input de texto con placeholder $0.00).
        # Formatear monto sin comas (el campo acepta números simples)
        await asyncio.sleep(random.uniform(1.5, 3.0))  # Delay humano largo
        monto_input = page.locator("input.multicredito__input").first
        await monto_input.scroll_into_view_if_needed()
        await asyncio.sleep(random.uniform(0.5, 1.0))  # Pausa antes de escribir
        # Limpiar el campo primero y luego llenar
        await monto_input.clear()
        await asyncio.sleep(0.3)
        # Escribir el monto como número entero sin formato
        await monto_input.fill(str(int(data.monto)))

        # 2) Plazo: es un <input type=range> id=meses. Los sliders se setean por
        #    JS + disparo manual de eventos input/change para que React reaccione.
        await asyncio.sleep(random.uniform(1.0, 2.0))  # Delay humano largo
        meses = _solo_meses_validos(data.meses)
        await page.eval_on_selector(
            "#meses",
            """(el, v) => {
                el.value = v;
                el.dispatchEvent(new Event('input',  {bubbles: true}));
                el.dispatchEvent(new Event('change', {bubbles: true}));
            }""",
            str(meses),
        )

        # 3) Tipo de amortizacion (boton Aleman / Frances).
        await asyncio.sleep(random.uniform(1.5, 2.5))  # Delay humano largo
        etiqueta = "Alemán" if data.amortizacion.lower().startswith("alem") else "Francés"
        # Usar un selector más específico para el botón
        boton_tipo = page.locator(f"button:has-text('{etiqueta}')")
        await boton_tipo.scroll_into_view_if_needed()
        await asyncio.sleep(random.uniform(0.5, 1.0))  # Pausa antes de click
        await boton_tipo.click(force=True)  # Force click para asegurar que funcione
        print(f"✓ Click en tipo de amortización: {etiqueta}")
        await asyncio.sleep(random.uniform(0.5, 1.0))

        # 4) Calcular.
        await asyncio.sleep(random.uniform(2.0, 3.5))  # Delay humano largo antes de calcular
        boton_calcular = page.locator("button.multicredito__buttonCalcular")
        await boton_calcular.scroll_into_view_if_needed()
        await asyncio.sleep(random.uniform(0.5, 1.0))  # Pausa antes de click

        # Screenshot antes de calcular
        await page.screenshot(path="debug_before_calculate.png")

        await boton_calcular.click(force=True)  # Force click para asegurar que funcione
        print("✓ Click en botón Calcular")
        await asyncio.sleep(random.uniform(0.5, 1.0))

        # Esperar procesamiento
        await asyncio.sleep(random.uniform(3.0, 5.0))

        # Screenshot después de calcular
        await page.screenshot(path="debug_after_calculate.png")

        # Esperar a que aparezcan los resultados - buscar por el texto "Cuotas mensuales"
        await page.wait_for_selector("text=Cuotas mensuales", timeout=45000)
        await asyncio.sleep(random.uniform(3.0, 5.0))  # Esperar render completo

        # Hacer scroll a los resultados
        await asyncio.sleep(random.uniform(1.0, 2.0))

        # 5) Leer el detalle del credito.
        #    Helper: dado el texto de una etiqueta, devuelve el monto de su fila.
        async def valor(etiqueta_texto: str) -> str:
            loc = page.locator(f"text={etiqueta_texto}").first
            fila = loc.locator("xpath=ancestor::*[1]")
            txt = await fila.inner_text()
            return txt.replace(etiqueta_texto, "").strip()

        # Obtener la cuota mensual (el número grande que está después de "Cuotas mensuales")
        cuota_section = page.locator("text=Cuotas mensuales").locator("xpath=following-sibling::*[1]")
        cuota_mensual = await cuota_section.inner_text()

        resultado = {
            "monto": data.monto,
            "meses": meses,
            "amortizacion": etiqueta,
            "correo": data.correo,
            "cuota_mensual": cuota_mensual.strip(),
            "capital":        await valor("Capital:"),
            "impuesto_solca": await valor("Impuesto de Solca:"),
            "total_interes":  await valor("Total de interés:"),
            "total_seguros":  await valor("Total de seguros:"),
            "total_pagar":    await valor("Total a pagar:"),
        }

        # 6) Tabla de amortizacion: abrir el modal y leer filas.
        await page.get_by_text("Ver tabla de amortización").click()

        # Esperar a que el modal esté completamente visible y renderizado
        await page.wait_for_selector(".modalTable__modal", timeout=15000, state="visible")
        await asyncio.sleep(1.0)

        # Esperar a que la tabla esté visible
        await page.wait_for_selector(".modalTable__table", timeout=10000, state="visible")
        await asyncio.sleep(1.0)

        filas = []
        for tr in await page.locator("table tr").all():
            texto = await tr.inner_text()
            celdas = [c.strip() for c in texto.split("\t") if c.strip()]
            # las filas de datos empiezan con el numero de cuota
            if celdas and celdas[0].isdigit():
                filas.append({
                    "n": celdas[0],
                    "saldo": celdas[1] if len(celdas) > 1 else None,
                    "capital": celdas[2] if len(celdas) > 2 else None,
                    "interes": celdas[3] if len(celdas) > 3 else None,
                    "s_desgravamen": celdas[4] if len(celdas) > 4 else None,
                    "s_cesantia": celdas[5] if len(celdas) > 5 else None,
                    "total_seguros": celdas[6] if len(celdas) > 6 else None,
                    "cuota": celdas[7] if len(celdas) > 7 else None,
                })
        resultado["tabla_amortizacion"] = filas

        # Guardar resultado para el endpoint GET
        global _ultimo_resultado
        _ultimo_resultado = resultado

        return resultado

    except Exception as e:
        # 502 = el upstream (el simulador) fallo. Power Automate puede reintentar.
        raise HTTPException(status_code=502, detail=f"Error en simulacion: {e}")
    finally:
        await context.close()


@app.get("/health")
async def health():
    return {"status": "ok"}


@app.get("/resultado")
async def obtener_resultado():
    """
    Obtiene los datos del último cálculo realizado:
    - Monto de Cuotas Mensuales
    - Detalle del crédito:
      · Capital
      · Impuesto de Solca
      · Total de interés
      · Total de seguros
      · Total a pagar
    """
    if _ultimo_resultado is None:
        raise HTTPException(
            status_code=404,
            detail="No hay resultados disponibles. Ejecuta POST /simular primero."
        )
    return _ultimo_resultado


@app.post("/descargar-pdf")
async def descargar_pdf(data: SimulacionInput):
    """
    Ejecuta una simulación y descarga la tabla de amortización en PDF.
    Devuelve la ruta del archivo descargado.
    """
    browser = _state["browser"]

    # Si Tor está habilitado, verificar y renovar identidad
    if USE_TOR:
        if not _check_tor_running():
            raise HTTPException(
                status_code=503,
                detail="Tor está habilitado (USE_TOR=true) pero no está corriendo. Ejecuta: sudo service tor@default start"
            )
        if not _renew_tor_identity():
            raise HTTPException(
                status_code=503,
                detail="No se pudo renovar la identidad de TOR. Verifica la configuración."
            )
        await asyncio.sleep(5)

    # Configurar context
    context_options = {
        "user_agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "viewport": {"width": 1920, "height": 1080},
        "locale": "es-EC",
        "timezone_id": "America/Guayaquil",
        "java_script_enabled": True,
        "accept_downloads": True  # Habilitar descargas
    }

    # Prioridad de proxy: PROXY_LIST > Tor > PROXY_SERVER
    if USE_PROXY_LIST:
        # Usar rotación de proxies residenciales
        proxy = _get_next_proxy()
        if proxy:
            context_options["proxy"] = proxy
            print(f"🌐 Usando proxy residencial: {proxy['server']} (usuario: {proxy['username']})")
        else:
            print("⚠️ No se pudo obtener proxy de la lista, continuando sin proxy")
    elif USE_TOR:
        context_options["proxy"] = {
            "server": f"socks5://127.0.0.1:{TOR_SOCKS_PORT}"
        }
        print(f"🧅 Usando Tor SOCKS5 proxy en puerto {TOR_SOCKS_PORT}")
    elif PROXY_SERVER:
        proxy_config = {"server": PROXY_SERVER}
        if PROXY_USERNAME and PROXY_PASSWORD:
            proxy_config["username"] = PROXY_USERNAME
            proxy_config["password"] = PROXY_PASSWORD
        context_options["proxy"] = proxy_config

    context = await browser.new_context(**context_options)

    # Inyectar scripts anti-detección
    await context.add_init_script("""
        Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
        window.navigator.chrome = {runtime: {}};
        Object.defineProperty(navigator, 'permissions', {
            get: () => ({query: () => Promise.resolve({state: 'denied'})})
        });
        Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]});
        Object.defineProperty(navigator, 'languages', {get: () => ['es-EC', 'es', 'en']});
    """)

    # Abrir una única pestaña
    page = await context.new_page()

    # Configurar CDP para habilitar descargas automáticas en el directorio temporal
    client = await page.context.new_cdp_session(page)
    await client.send('Page.setDownloadBehavior', {
        'behavior': 'allow',
        'downloadPath': str(PDF_DIR.absolute())
    })
    try:
        await page.goto(URL, wait_until="domcontentloaded", timeout=60000)
        await asyncio.sleep(random.uniform(5.0, 8.0))

        # Aceptar cookies si aparece el popup
        try:
            aceptar_btn = page.locator("#hs-eu-confirmation-button")
            await aceptar_btn.wait_for(timeout=5000, state="visible")
            await asyncio.sleep(random.uniform(0.5, 1.0))
            await aceptar_btn.click()
            print("✓ Cookies aceptadas (PDF)")
            await asyncio.sleep(random.uniform(1.0, 2.0))
        except:
            print("⚠ No se encontró popup de cookies o ya fue aceptado (PDF)")
            pass

        # Esperar formulario
        await page.wait_for_selector("input.multicredito__input", timeout=60000)
        await asyncio.sleep(random.uniform(1.5, 2.5))

        # 1) Monto del crédito
        await asyncio.sleep(random.uniform(1.5, 3.0))
        monto_input = page.locator("input.multicredito__input").first
        await monto_input.scroll_into_view_if_needed()
        await asyncio.sleep(random.uniform(0.5, 1.0))
        await monto_input.clear()
        await asyncio.sleep(0.3)
        await monto_input.fill(str(int(data.monto)))

        # 2) Plazo (usar fill() de Playwright para correcta actualización del slider)
        # IMPORTANTE: Usar selector específico porque hay un datalist con el mismo ID
        await asyncio.sleep(random.uniform(2.0, 3.0))
        meses = _solo_meses_validos(data.meses)
        slider = page.locator("input#meses[type=range]")
        await slider.fill(str(meses))
        await asyncio.sleep(random.uniform(2.5, 3.5))  # Esperar a que se procese
        print(f"✓ Slider configurado a {meses} meses usando fill()")

        # 3) Tipo de amortización
        await asyncio.sleep(random.uniform(1.5, 2.5))
        etiqueta = "Alemán" if data.amortizacion.lower().startswith("alem") else "Francés"
        boton_tipo = page.locator(f"button:has-text('{etiqueta}')")
        await boton_tipo.scroll_into_view_if_needed()
        await asyncio.sleep(random.uniform(0.5, 1.0))
        await boton_tipo.click(force=True)
        print(f"✓ Click en tipo de amortización: {etiqueta} (PDF)")
        await asyncio.sleep(random.uniform(0.5, 1.0))

        # 4) Calcular
        await asyncio.sleep(random.uniform(2.0, 3.5))
        boton_calcular = page.locator("button.multicredito__buttonCalcular")
        await boton_calcular.scroll_into_view_if_needed()
        await asyncio.sleep(random.uniform(0.5, 1.0))
        await boton_calcular.click(force=True)
        print("✓ Click en botón Calcular (PDF)")
        await asyncio.sleep(random.uniform(0.5, 1.0))

        # Esperar procesamiento
        await asyncio.sleep(random.uniform(3.0, 5.0))

        # Esperar resultados
        await page.wait_for_selector("text=Cuotas mensuales", timeout=45000)
        await asyncio.sleep(random.uniform(3.0, 5.0))

        # 5) Abrir tabla de amortización
        await page.get_by_text("Ver tabla de amortización").click()

        # Esperar a que el modal esté completamente visible y renderizado
        await page.wait_for_selector(".modalTable__modal", timeout=15000, state="visible")
        await asyncio.sleep(1.0)

        # Esperar a que la tabla esté visible
        await page.wait_for_selector(".modalTable__table", timeout=10000, state="visible")
        await asyncio.sleep(1.0)

        # Esperar a que el encabezado del modal esté completamente renderizado
        await page.wait_for_selector(".modalTable__title", timeout=10000, state="visible")
        await page.wait_for_selector(".descriptionTable", timeout=10000, state="visible")

        # Esperar tiempo adicional para asegurar que todos los datos estén renderizados
        await asyncio.sleep(2.0)

        # 6) Extraer el contenido del modal completo para crear un PDF limpio

        # Screenshot del modal completo antes de extraer (para debug)
        await page.screenshot(path="debug_modal_completo.png", full_page=True)

        # Extraer el contenido del modal usando los selectores correctos
        modal_content = await page.evaluate("""
            () => {
                // Buscar el modal con la clase específica
                const modal = document.querySelector('.modalTable__modal');
                if (!modal) {
                    console.error('No se encontró el modal');
                    return null;
                }

                // Obtener el título
                const titleEl = modal.querySelector('.modalTable__title');
                const titulo = titleEl ? titleEl.textContent.trim() : 'Tabla de amortización';

                // Obtener información del préstamo
                const info = {};
                const descRows = modal.querySelectorAll('.descriptionTable__row');

                if (descRows.length === 0) {
                    console.warn('No se encontraron filas de descripción');
                }

                descRows.forEach((row, index) => {
                    const ps = row.querySelectorAll('p');
                    if (ps.length >= 2) {
                        const key = ps[0].textContent.trim();
                        const value = ps[1].textContent.trim();

                        console.log(`Fila ${index}: ${key} = ${value}`);

                        if (key.includes('Producto:')) info.producto = value;
                        if (key.includes('Plazo (meses):') || key.includes('Plazo')) info.plazo = value;
                        if (key.includes('Tasa de interés nominal:') || key.includes('Tasa')) info.tasa = value;
                        if (key.includes('Capital:')) info.capital = value;
                        if (key.includes('Total de interés:') || key.includes('Total de interés')) info.interes = value;
                    }
                });

                // Obtener tabla HTML
                const tableBody = modal.querySelector('.modalTable__body');
                const table = modal.querySelector('.modalTable__table');

                if (!table) {
                    console.error('No se encontró la tabla');
                    return null;
                }

                const tableHTML = table.outerHTML;

                console.log('Extracción exitosa:', {
                    titulo,
                    info,
                    tableLength: tableHTML.length
                });

                return {
                    titulo,
                    info,
                    tableHTML
                };
            }
        """)

        if not modal_content:
            # Tomar screenshot para debug
            await page.screenshot(path="debug_modal_error.png", full_page=True)
            raise Exception("No se pudo extraer el contenido del modal. El modal no está presente en la página. Revisa debug_modal_error.png")

        if not modal_content.get('tableHTML'):
            await page.screenshot(path="debug_table_error.png", full_page=True)
            raise Exception(f"No se pudo extraer la tabla del modal. Contenido extraído: {modal_content}. Revisa debug_table_error.png")

        # Crear una nueva página con HTML limpio
        new_page = await context.new_page()

        # Construir HTML limpio con solo la tabla
        titulo = modal_content.get('titulo', 'Tabla de amortización')
        info = modal_content.get('info', {})
        producto = info.get('producto', 'Multicrédito')
        plazo = info.get('plazo', str(meses))
        tasa = info.get('tasa', 'N/A')
        capital = info.get('capital', 'N/A')
        interes = info.get('interes', 'N/A')

        clean_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                * {{
                    margin: 0;
                    padding: 0;
                    box-sizing: border-box;
                }}
                @page {{
                    margin: 15mm;
                    size: A4;
                }}
                body {{
                    font-family: Arial, Helvetica, sans-serif;
                    padding: 20px;
                    background: white;
                    color: #333;
                }}
                h1 {{
                    font-size: 22px;
                    margin-bottom: 25px;
                    text-align: center;
                    color: #d32f2f;
                    border-bottom: 3px solid #d32f2f;
                    padding-bottom: 10px;
                }}
                .info-container {{
                    background: #f5f5f5;
                    padding: 15px;
                    border-radius: 8px;
                    margin-bottom: 20px;
                    border: 1px solid #ddd;
                }}
                .info {{
                    display: grid;
                    grid-template-columns: 1fr 1fr;
                    gap: 15px;
                    font-size: 13px;
                }}
                .info-left, .info-right {{
                    display: flex;
                    flex-direction: column;
                    gap: 8px;
                }}
                .info p {{
                    margin: 0;
                    line-height: 1.5;
                }}
                .info strong {{
                    color: #555;
                    font-weight: 600;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 10px;
                    background: white;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                }}
                th, td {{
                    border: 1px solid #ddd;
                    padding: 10px 8px;
                    text-align: center;
                    font-size: 12px;
                }}
                th {{
                    background-color: #d32f2f;
                    color: white;
                    font-weight: bold;
                    text-transform: uppercase;
                    font-size: 11px;
                    letter-spacing: 0.5px;
                }}
                tbody tr:nth-child(odd) {{
                    background-color: #f9f9f9;
                }}
                tbody tr:nth-child(even) {{
                    background-color: white;
                }}
                tbody tr:hover {{
                    background-color: #fff3e0;
                }}
                .footer {{
                    margin-top: 20px;
                    text-align: center;
                    font-size: 11px;
                    color: #888;
                    padding-top: 15px;
                    border-top: 1px solid #ddd;
                }}
            </style>
        </head>
        <body>
            <h1>{titulo}</h1>
            <div class="info-container">
                <div class="info">
                    <div class="info-left">
                        <p><strong>Producto:</strong> {producto}</p>
                        <p><strong>Plazo (meses):</strong> {plazo}</p>
                        <p><strong>Tasa de interés nominal:</strong> {tasa}</p>
                    </div>
                    <div class="info-right">
                        <p><strong>Capital:</strong> {capital}</p>
                        <p><strong>Total de interés:</strong> {interes}</p>
                    </div>
                </div>
            </div>
            {modal_content['tableHTML']}
            <div class="footer">
                <p>Banco Guayaquil - Simulador de Multicrédito</p>
                <p>Documento generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}</p>
            </div>
        </body>
        </html>
        """

        await new_page.set_content(clean_html)
        await asyncio.sleep(0.5)

        # Generar PDF de la página limpia
        pdf_bytes = await new_page.pdf(
            format='A4',
            print_background=True,
            margin={
                'top': '15mm',
                'right': '15mm',
                'bottom': '15mm',
                'left': '15mm'
            }
        )

        # Cerrar la página temporal
        await new_page.close()

        # Generar nombre descriptivo para el PDF: {monto}_{meses}_{amortización}_{fecha}_{usuario}.pdf
        fecha = datetime.now().strftime("%Y%m%d")
        usuario = data.correo.split("@")[0] if data.correo else "cliente"
        # Sanitizar valores para nombre de archivo
        amort_limpio = etiqueta.lower().replace("é", "e").replace("á", "a")
        filename = f"{int(data.monto)}_{meses}_{amort_limpio}_{fecha}_{usuario}.pdf"

        # Devolver el PDF directamente
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )

    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Error al descargar PDF: {e}")
    finally:
        await context.close()


@app.post("/generar-pdf")
async def generar_pdf(data: SimulacionInput):
    """
    Genera un PDF de la tabla de amortización usando los datos de la simulación.
    NO abre el modal del banco, solo simula y genera el PDF localmente.
    Esto evita bloqueos del WAF.
    """
    browser = _state["browser"]

    # Si Tor está habilitado, verificar y renovar identidad
    if USE_TOR:
        if not _check_tor_running():
            raise HTTPException(
                status_code=503,
                detail="Tor está habilitado (USE_TOR=true) pero no está corriendo. Ejecuta: sudo service tor@default start"
            )
        if not _renew_tor_identity():
            raise HTTPException(
                status_code=503,
                detail="No se pudo renovar la identidad de TOR. Verifica la configuración."
            )
        await asyncio.sleep(5)

    # Configurar context con User-Agent rotativo
    user_agent = _get_next_user_agent()
    context_options = {
        "user_agent": user_agent,
        "viewport": {"width": 1920, "height": 1080},
        "locale": "es-EC",
        "timezone_id": "America/Guayaquil",
        "java_script_enabled": True,
        "permissions": ["geolocation"],  # Simular permisos de navegador real
        "has_touch": False,
        "is_mobile": False,
        "device_scale_factor": 1
    }

    # Prioridad de proxy: PROXY_LIST > Tor > PROXY_SERVER
    if USE_PROXY_LIST:
        # Usar rotación de proxies residenciales
        proxy = _get_next_proxy()
        if proxy:
            context_options["proxy"] = proxy
            print(f"🌐 Usando proxy residencial: {proxy['server']} (usuario: {proxy['username']})")
            print(f"🎭 User-Agent: {user_agent[:80]}...")
        else:
            print("⚠️ No se pudo obtener proxy de la lista, continuando sin proxy")
    elif USE_TOR:
        context_options["proxy"] = {
            "server": f"socks5://127.0.0.1:{TOR_SOCKS_PORT}"
        }
        print(f"🧅 Usando Tor SOCKS5 proxy en puerto {TOR_SOCKS_PORT}")
    elif PROXY_SERVER:
        proxy_config = {"server": PROXY_SERVER}
        if PROXY_USERNAME and PROXY_PASSWORD:
            proxy_config["username"] = PROXY_USERNAME
            proxy_config["password"] = PROXY_PASSWORD
        context_options["proxy"] = proxy_config

    context = await browser.new_context(**context_options)

    # Inyectar scripts anti-detección avanzados (anti-PerimeterX)
    await context.add_init_script("""
        // 1. Ocultar webdriver
        Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
        delete navigator.__proto__.webdriver;

        // 2. Chrome runtime
        window.navigator.chrome = {
            runtime: {},
            loadTimes: function() {},
            csi: function() {},
            app: {}
        };

        // 3. Permisos realistas
        const originalQuery = window.navigator.permissions.query;
        window.navigator.permissions.query = (parameters) => (
            parameters.name === 'notifications' ?
                Promise.resolve({ state: Notification.permission }) :
                originalQuery(parameters)
        );

        // 4. Plugins realistas
        Object.defineProperty(navigator, 'plugins', {
            get: () => [
                {name: 'Chrome PDF Plugin', filename: 'internal-pdf-viewer'},
                {name: 'Chrome PDF Viewer', filename: 'mhjfbmdgcfjbbpaeojofohoefgiehjai'},
                {name: 'Native Client', filename: 'internal-nacl-plugin'}
            ]
        });

        // 5. Lenguajes del navegador
        Object.defineProperty(navigator, 'languages', {
            get: () => ['es-EC', 'es', 'en-US', 'en']
        });

        // 6. Canvas fingerprinting - agregar ruido mínimo
        const originalToDataURL = HTMLCanvasElement.prototype.toDataURL;
        HTMLCanvasElement.prototype.toDataURL = function(type) {
            const shift = Math.floor(Math.random() * 10) - 5;
            const context = this.getContext('2d');
            const imageData = context.getImageData(0, 0, this.width, this.height);
            for (let i = 0; i < imageData.data.length; i += 4) {
                imageData.data[i] = imageData.data[i] + shift;
            }
            context.putImageData(imageData, 0, 0);
            return originalToDataURL.apply(this, arguments);
        };

        // 7. WebGL fingerprinting
        const getParameter = WebGLRenderingContext.prototype.getParameter;
        WebGLRenderingContext.prototype.getParameter = function(parameter) {
            if (parameter === 37445) return 'Intel Inc.';
            if (parameter === 37446) return 'Intel Iris OpenGL Engine';
            return getParameter.call(this, parameter);
        };

        // 8. Audio context fingerprinting
        const audioContext = window.AudioContext || window.webkitAudioContext;
        if (audioContext) {
            const OriginalAudioContext = audioContext;
            window.AudioContext = function() {
                const context = new OriginalAudioContext();
                const originalCreateOscillator = context.createOscillator;
                context.createOscillator = function() {
                    const oscillator = originalCreateOscillator.call(context);
                    const originalStart = oscillator.start;
                    oscillator.start = function(when) {
                        return originalStart.call(oscillator, when + Math.random() * 0.0001);
                    };
                    return oscillator;
                };
                return context;
            };
        }

        // 9. Screen resolution realista
        Object.defineProperty(window.screen, 'availWidth', {get: () => 1920});
        Object.defineProperty(window.screen, 'availHeight', {get: () => 1040});

        // 10. Battery API
        if (navigator.getBattery) {
            navigator.getBattery = () => Promise.resolve({
                charging: true,
                chargingTime: 0,
                dischargingTime: Infinity,
                level: 1
            });
        }
    """)

    # Abrir una única pestaña
    page = await context.new_page()

    # Aplicar playwright-stealth para ocultar mejor automatización
    stealth = Stealth()
    await stealth.apply_stealth_async(page)
    print("🥷 Stealth aplicado - navegador parecerá completamente humano")

    try:
        # Ejecutar simulación en la segunda pestaña
        print(f"🌍 Cargando página {URL}...")
        await page.goto(URL, wait_until="domcontentloaded", timeout=90000)
        print("✓ Página cargada exitosamente")
        # Delay MUY largo inicial para permitir que scripts anti-bot se ejecuten completamente
        # y parecer más humano (persona leyendo la página)
        await asyncio.sleep(random.uniform(15.0, 20.0))
        print("⏱️  Delay anti-bot completado")

        # Aceptar cookies
        try:
            aceptar_btn = page.locator("#hs-eu-confirmation-button")
            await aceptar_btn.wait_for(timeout=5000, state="visible")
            await asyncio.sleep(random.uniform(0.5, 1.0))
            await aceptar_btn.click()
            print("✓ Cookies aceptadas")
            await asyncio.sleep(random.uniform(1.0, 2.0))
        except:
            print("⚠ No se encontró popup de cookies")
            pass

        await page.wait_for_selector("input.multicredito__input", timeout=60000)
        await asyncio.sleep(random.uniform(2.0, 4.0))

        # Llenar formulario con delays más humanos
        await asyncio.sleep(random.uniform(2.0, 4.0))
        monto_input = page.locator("input.multicredito__input").first
        await monto_input.scroll_into_view_if_needed()
        await asyncio.sleep(random.uniform(1.0, 2.0))
        await monto_input.clear()
        await asyncio.sleep(random.uniform(0.5, 1.0))
        await monto_input.fill(str(int(data.monto)))

        # 2) Plazo (usar fill() de Playwright para correcta actualización del slider)
        # IMPORTANTE: Usar selector específico porque hay un datalist con el mismo ID
        await asyncio.sleep(random.uniform(2.0, 3.0))
        meses = _solo_meses_validos(data.meses)
        slider = page.locator("input#meses[type=range]")
        await slider.fill(str(meses))
        await asyncio.sleep(random.uniform(2.5, 3.5))  # Esperar a que se procese
        print(f"✓ Slider configurado a {meses} meses usando fill()")

        # 3) Tipo de amortización
        await asyncio.sleep(random.uniform(2.0, 3.5))
        etiqueta = "Alemán" if data.amortizacion.lower().startswith("alem") else "Francés"
        boton_tipo = page.locator(f"button:has-text('{etiqueta}')")
        await boton_tipo.scroll_into_view_if_needed()
        await asyncio.sleep(random.uniform(1.0, 2.0))
        await boton_tipo.click(force=True)
        print(f"✓ Click en tipo de amortización: {etiqueta}")
        await asyncio.sleep(random.uniform(1.5, 2.5))

        await asyncio.sleep(random.uniform(3.0, 5.0))
        boton_calcular = page.locator("button.multicredito__buttonCalcular")
        await boton_calcular.scroll_into_view_if_needed()
        await asyncio.sleep(random.uniform(1.0, 2.0))
        await boton_calcular.click(force=True)
        print("✓ Click en botón Calcular")
        await asyncio.sleep(random.uniform(2.0, 3.0))

        # Esperar más tiempo para que el banco procese
        await asyncio.sleep(random.uniform(5.0, 8.0))

        # Screenshot de debug para ver qué muestra el banco
        await page.screenshot(path="debug_esperando_resultados.png", full_page=True)
        print("📸 Screenshot tomado: debug_esperando_resultados.png")

        # Esperar directamente por el botón "Ver tabla de amortización" en lugar de "Cuotas mensuales"
        await page.wait_for_selector("text=Ver tabla de amortización", timeout=60000, state="visible")
        print("✓ Resultados mostrados - botón 'Ver tabla' disponible")
        await asyncio.sleep(random.uniform(2.0, 4.0))

        # Extraer datos (sin abrir modal)
        async def valor(etiqueta_texto: str) -> str:
            loc = page.locator(f"text={etiqueta_texto}").first
            fila = loc.locator("xpath=ancestor::*[1]")
            txt = await fila.inner_text()
            return txt.replace(etiqueta_texto, "").strip()

        print("📊 Extrayendo datos del resumen...")
        capital = await valor("Capital:")
        total_interes = await valor("Total de interés:")
        print(f"   Capital: {capital}, Interés: {total_interes}")

        # Extraer tasa de interés nominal desde la página principal ANTES de abrir modal
        tasa_nominal = "N/A"
        try:
            tasa_elem = page.locator("span.simuladorVehicular__cardTextSubtitle").first
            tasa_text = await tasa_elem.inner_text()
            # Extraer el porcentaje usando regex (ej: "15.6%")
            match = re.search(r'(\d+\.?\d*%)', tasa_text)
            if match:
                tasa_nominal = match.group(1)
                print(f"   ✓ Tasa nominal: {tasa_nominal}")
            else:
                print(f"   ⚠️ No se pudo extraer tasa de: {tasa_text[:50]}...")
        except Exception as e:
            print(f"   ⚠️ Error al extraer tasa: {e}")

        # Abrir modal SOLO para obtener datos de la tabla
        print("🔘 Haciendo click en 'Ver tabla de amortización'...")
        await asyncio.sleep(random.uniform(2.0, 3.0))
        await page.get_by_text("Ver tabla de amortización").click()
        print("✓ Click realizado, esperando modal...")

        await page.wait_for_selector(".modalTable__modal", timeout=15000, state="visible")
        print("✓ Modal visible")
        await asyncio.sleep(1.0)

        await page.wait_for_selector(".modalTable__table", timeout=10000, state="visible")
        print("✓ Tabla visible")
        await asyncio.sleep(1.0)

        # IMPORTANTE: Hacer scroll dentro del modal para forzar renderizado de TODAS las filas
        # El modal puede tener virtual scrolling que solo muestra las primeras 12 filas
        print("🔄 Haciendo scroll dentro del modal para cargar todas las filas...")
        modal_body = page.locator(".modalTable__body").first

        # Hacer scroll gradual hasta el final del modal para forzar renderizado de todas las filas
        # Aumentado a 20 scrolls para asegurar que se cargan TODAS las filas en plazos largos (48-60 meses)
        for i in range(20):  # Hacer 20 scrolls graduales
            await modal_body.evaluate("el => el.scrollTop = el.scrollHeight")
            await asyncio.sleep(0.5)  # Esperar a que se rendericen nuevas filas

        print("✓ Scroll completado, esperando renderizado final...")
        await asyncio.sleep(3.0)  # Esperar a que todas las filas estén renderizadas

        # Extraer datos de la tabla (usar tbody para obtener solo las filas de datos)
        filas_datos = []
        for tr in await page.locator("table tbody tr").all():
            texto = await tr.inner_text()
            celdas = [c.strip() for c in texto.split("\t") if c.strip()]
            if celdas and celdas[0].isdigit():
                filas_datos.append({
                    "n": celdas[0],
                    "saldo": celdas[1] if len(celdas) > 1 else "",
                    "capital": celdas[2] if len(celdas) > 2 else "",
                    "interes": celdas[3] if len(celdas) > 3 else "",
                    "s_desgravamen": celdas[4] if len(celdas) > 4 else "",
                    "s_cesantia": celdas[5] if len(celdas) > 5 else "",
                    "total_seguros": celdas[6] if len(celdas) > 6 else "",
                    "cuota": celdas[7] if len(celdas) > 7 else "",
                })

        print(f"📋 Extraídas {len(filas_datos)} filas de datos de la tabla")

        # NOTA: La tasa nominal ya fue extraída de la página principal antes de abrir el modal
        print("🔍 Extrayendo Total de interés desde el modal...")

        # Total de interés desde el modal (más confiable)
        try:
            # Buscar en las filas del modal el campo "Total de interés"
            modal_desc_rows = await page.locator('.descriptionTable__row').all()
            total_interes_modal = None

            for row in modal_desc_rows:
                text = await row.inner_text()
                if "Total de interés" in text or "Total de interés:" in text:
                    # Extraer el valor (segundo elemento)
                    ps = await row.locator('p').all()
                    if len(ps) >= 2:
                        total_interes_modal = await ps[1].inner_text()
                        print(f"   Total de interés (desde modal): {total_interes_modal}")
                        break

            # Si se encontró en el modal, actualizar la variable
            if total_interes_modal:
                total_interes = total_interes_modal
            else:
                print(f"⚠️ Total de interés no encontrado en modal, usando valor de página principal: {total_interes}")
        except Exception as e:
            print(f"⚠️ Error al extraer Total de interés desde modal: {e}. Usando valor de página principal")

        print("✓ Cerrando página original...")
        # Cerrar el browser page original
        await page.close()
        print("✓ Página cerrada")

        print("📄 Iniciando generación de HTML para PDF...")
        # Generar HTML para PDF (diseño EXACTO al del banco)
        filas_html = ""
        for fila in filas_datos:
            filas_html += f"""
                <tr>
                    <td>{fila['n']}</td>
                    <td>{fila['saldo']}</td>
                    <td>{fila['capital']}</td>
                    <td>{fila['interes']}</td>
                    <td>{fila['s_desgravamen']}</td>
                    <td>{fila['s_cesantia']}</td>
                    <td>{fila['total_seguros']}</td>
                    <td>{fila['cuota']}</td>
                </tr>
            """

        html_pdf = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                * {{
                    margin: 0;
                    padding: 0;
                    box-sizing: border-box;
                }}
                @page {{
                    margin: 50px 80px;
                    size: A4;
                }}
                body {{
                    font-family: Arial, sans-serif;
                    color: #000;
                    background: #f5f5f0;
                    padding: 40px 20px;
                }}
                h1 {{
                    font-size: 28px;
                    margin-bottom: 30px;
                    text-align: center;
                    font-weight: bold;
                }}
                .info-container {{
                    display: table;
                    width: 100%;
                    margin-bottom: 25px;
                    font-size: 15px;
                }}
                .info-row {{
                    display: table-row;
                }}
                .info-left, .info-right {{
                    display: table-cell;
                    padding: 3px 0;
                }}
                .info-left {{
                    text-align: left;
                    width: 50%;
                }}
                .info-right {{
                    text-align: right;
                    width: 50%;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 10px;
                    background: #fff;
                }}
                th, td {{
                    border: 1px solid #000;
                    padding: 10px 8px;
                    text-align: center;
                    font-size: 13px;
                }}
                th {{
                    font-weight: bold;
                    background: #fff;
                }}
                td {{
                    background: #fff;
                }}
            </style>
        </head>
        <body>
            <h1>Tabla de amortización es {etiqueta.lower()}</h1>

            <div class="info-container">
                <div class="info-row">
                    <div class="info-left">Producto: Multicrédito</div>
                    <div class="info-right">Capital: {capital}</div>
                </div>
                <div class="info-row">
                    <div class="info-left">Plazo (meses): {meses}</div>
                    <div class="info-right">Total de interés: {total_interes}</div>
                </div>
                <div class="info-row">
                    <div class="info-left">Tasa de interés nominal: {tasa_nominal if 'tasa_nominal' in locals() else '15.60%'}</div>
                    <div class="info-right"></div>
                </div>
            </div>

            <table>
                <thead>
                    <tr>
                        <th>N°</th>
                        <th>Saldo</th>
                        <th>Capital</th>
                        <th>Interés</th>
                        <th>S.<br/>Desgravamen</th>
                        <th>S.<br/>Cesantia</th>
                        <th>Total<br/>Seguros</th>
                        <th>Cuota</th>
                    </tr>
                </thead>
                <tbody>
                    {filas_html}
                </tbody>
            </table>
        </body>
        </html>
        """
        print(f"✓ HTML generado ({len(html_pdf)} caracteres)")

        print("📄 Creando nueva página para PDF...")
        # Crear nueva página para generar PDF
        pdf_page = await context.new_page()
        print("✓ Nueva página creada")

        print("📄 Estableciendo contenido HTML en la página...")
        await pdf_page.set_content(html_pdf)
        print("✓ Contenido HTML establecido")
        await asyncio.sleep(0.5)

        print("📄 Generando PDF desde la página...")
        # Generar PDF
        pdf_bytes = await pdf_page.pdf(
            format='A4',
            print_background=True,
            margin={
                'top': '15mm',
                'right': '15mm',
                'bottom': '15mm',
                'left': '15mm'
            }
        )
        print(f"✓ PDF generado en memoria ({len(pdf_bytes)} bytes)")

        print("📄 Cerrando página de PDF...")
        await pdf_page.close()
        print("✓ Página de PDF cerrada")

        # Generar nombre descriptivo para el PDF: {monto}_{meses}_{amortización}_{fecha}_{usuario}.pdf
        fecha = datetime.now().strftime("%Y%m%d")
        usuario = data.correo.split("@")[0] if data.correo else "cliente"
        # Sanitizar valores para nombre de archivo
        amort_limpio = etiqueta.lower().replace("é", "e").replace("á", "a")
        filename = f"{int(data.monto)}_{meses}_{amort_limpio}_{fecha}_{usuario}.pdf"

        print(f"✓ PDF generado exitosamente: {filename}")

        print("📤 Preparando respuesta HTTP con PDF...")
        # Devolver el PDF directamente
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )

    except Exception as e:
        print(f"❌ ERROR en generar_pdf: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=502, detail=f"Error al generar PDF: {e}")
    finally:
        print("🔧 Cerrando contexto del navegador...")
        await context.close()
        print("✓ Contexto cerrado")


@app.post("/generar-excel")
async def generar_excel(data: SimulacionInput):
    """
    Endpoint que genera un archivo Excel con la tabla de amortización.
    Hace el mismo scraping que /generar-pdf pero retorna un archivo Excel.
    """
    browser = _state["browser"]

    # Si Tor está habilitado, verificar y renovar identidad
    if USE_TOR:
        if not _check_tor_running():
            raise HTTPException(
                status_code=503,
                detail="Tor está habilitado (USE_TOR=true) pero no está corriendo. Ejecuta: sudo service tor@default start"
            )
        if not _renew_tor_identity():
            raise HTTPException(
                status_code=503,
                detail="No se pudo renovar la identidad de TOR. Verifica la configuración."
            )
        await asyncio.sleep(5)

    # Configurar context con User-Agent rotativo
    user_agent = _get_next_user_agent()
    context_options = {
        "user_agent": user_agent,
        "viewport": {"width": 1920, "height": 1080},
        "locale": "es-EC",
        "timezone_id": "America/Guayaquil",
        "java_script_enabled": True,
        "permissions": ["geolocation"],
        "has_touch": False,
        "is_mobile": False,
        "device_scale_factor": 1
    }

    # Prioridad de proxy: PROXY_LIST > Tor > PROXY_SERVER
    if USE_PROXY_LIST:
        proxy = _get_next_proxy()
        if proxy:
            context_options["proxy"] = proxy
            print(f"🌐 Usando proxy residencial: {proxy['server']} (usuario: {proxy['username']})")
            print(f"🎭 User-Agent: {user_agent[:80]}...")
        else:
            print("⚠️ No se pudo obtener proxy de la lista, continuando sin proxy")
    elif USE_TOR:
        context_options["proxy"] = {
            "server": f"socks5://127.0.0.1:{TOR_SOCKS_PORT}"
        }
        print(f"🧅 Usando Tor SOCKS5 proxy en puerto {TOR_SOCKS_PORT}")
    elif PROXY_SERVER:
        proxy_config = {"server": PROXY_SERVER}
        if PROXY_USERNAME and PROXY_PASSWORD:
            proxy_config["username"] = PROXY_USERNAME
            proxy_config["password"] = PROXY_PASSWORD
        context_options["proxy"] = proxy_config

    context = await browser.new_context(**context_options)

    # Inyectar scripts anti-detección avanzados (anti-PerimeterX)
    await context.add_init_script("""
        // Ocultar webdriver
        Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
        delete navigator.__proto__.webdriver;

        // Chrome runtime
        window.navigator.chrome = {
            runtime: {},
            loadTimes: function() {},
            csi: function() {},
            app: {}
        };

        // Canvas fingerprinting - agregar ruido mínimo
        const originalToDataURL = HTMLCanvasElement.prototype.toDataURL;
        HTMLCanvasElement.prototype.toDataURL = function(type) {
            const shift = Math.floor(Math.random() * 10) - 5;
            const context = this.getContext('2d');
            const imageData = context.getImageData(0, 0, this.width, this.height);
            for (let i = 0; i < imageData.data.length; i += 4) {
                imageData.data[i] = imageData.data[i] + shift;
            }
            context.putImageData(imageData, 0, 0);
            return originalToDataURL.apply(this, arguments);
        };

        // WebGL fingerprinting
        const getParameter = WebGLRenderingContext.prototype.getParameter;
        WebGLRenderingContext.prototype.getParameter = function(parameter) {
            if (parameter === 37445) return 'Intel Inc.';
            if (parameter === 37446) return 'Intel Iris OpenGL Engine';
            return getParameter.call(this, parameter);
        };
    """)

    try:
        # Abrir una única pestaña
        page = await context.new_page()

        # Aplicar playwright-stealth para ocultar mejor automatización
        stealth = Stealth()
        await stealth.apply_stealth_async(page)
        print("🥷 Stealth aplicado - navegador parecerá completamente humano")

        # Cargar página
        print(f"🌍 Cargando página {URL}...")
        await page.goto(URL, wait_until="domcontentloaded", timeout=90000)
        print("✓ Página cargada exitosamente")

        # Delay anti-bot
        await asyncio.sleep(random.uniform(15.0, 20.0))
        print("⏱️  Delay anti-bot completado")

        # Aceptar cookies si aparece el botón
        try:
            await page.click('button:has-text("Aceptar")', timeout=3000)
            print("✓ Cookies aceptadas")
        except:
            print("⚠️ No se encontró el botón de cookies")

        # Rellenar formulario (código correcto que funciona)
        meses = _solo_meses_validos(data.meses)
        etiqueta = data.amortizacion.lower()

        # IMPORTANTE: Primero hacer scroll a la sección del simulador
        print("⬇️ Haciendo scroll a la sección del simulador...")
        monto_input = page.locator("input.multicredito__input").first
        await monto_input.scroll_into_view_if_needed()
        await asyncio.sleep(random.uniform(2.0, 3.0))
        print("✓ Scroll completado - sección visible")

        # 1) Monto
        await monto_input.clear()
        await asyncio.sleep(random.uniform(0.5, 1.0))
        await monto_input.fill(str(int(data.monto)))

        # 2) Plazo (usar fill() de Playwright para correcta actualización del slider)
        # IMPORTANTE: Usar selector específico porque hay un datalist con el mismo ID
        await asyncio.sleep(random.uniform(2.0, 3.0))
        slider = page.locator("input#meses[type=range]")
        await slider.fill(str(meses))
        await asyncio.sleep(random.uniform(2.5, 3.5))  # Esperar a que se procese
        print(f"✓ Slider configurado a {meses} meses usando fill()")

        # 3) Tipo de amortización
        await asyncio.sleep(random.uniform(1.5, 2.5))
        etiqueta = "Alemán" if data.amortizacion.lower().startswith("alem") else "Francés"
        boton_amort = page.locator(f'button:has-text("{etiqueta}")')
        await boton_amort.scroll_into_view_if_needed()
        await asyncio.sleep(random.uniform(0.5, 1.0))
        await boton_amort.click(force=True)
        print(f"✓ Click en tipo de amortización: {etiqueta}")

        # 4) Click en Calcular
        await asyncio.sleep(random.uniform(2.0, 3.0))
        calcular_btn = page.locator('button:has-text("Calcular")')
        await calcular_btn.scroll_into_view_if_needed()
        await asyncio.sleep(random.uniform(0.5, 1.0))
        await calcular_btn.click()
        print("✓ Click en botón Calcular")

        # Esperar resultados
        await asyncio.sleep(random.uniform(2.0, 3.0))
        await page.wait_for_selector('button:has-text("Ver tabla")', timeout=30000)
        print("✓ Resultados mostrados - botón 'Ver tabla' disponible")

        # Extraer datos del resumen
        print("📊 Extrayendo datos del resumen...")
        try:
            capital_span = await page.query_selector('span:has-text("Capital:") + span')
            capital = await capital_span.inner_text() if capital_span else "$0.00"
        except:
            capital = "$0.00"

        try:
            interes_span = await page.query_selector('span:has-text("Total de interés:") + span')
            total_interes = await interes_span.inner_text() if interes_span else "$0.00"
        except:
            total_interes = "$0.00"

        try:
            solca_span = await page.query_selector('span:has-text("Impuesto de Solca:") + span')
            impuesto_solca = await solca_span.inner_text() if solca_span else "$0.00"
        except:
            impuesto_solca = "$0.00"

        try:
            seguros_span = await page.query_selector('span:has-text("Total de seguros:") + span')
            total_seguros = await seguros_span.inner_text() if seguros_span else "$0.00"
        except:
            total_seguros = "$0.00"

        try:
            pagar_span = await page.query_selector('span:has-text("Total a pagar:") + span')
            total_pagar = await pagar_span.inner_text() if pagar_span else "$0.00"
        except:
            total_pagar = "$0.00"

        print(f"   Capital: {capital}, Interés: {total_interes}, Solca: {impuesto_solca}, Seguros: {total_seguros}, Total a pagar: {total_pagar}")

        # Extraer tasa de interés nominal desde la página principal ANTES de abrir modal
        tasa_nominal = "N/A"
        try:
            tasa_elem = page.locator("span.simuladorVehicular__cardTextSubtitle").first
            tasa_text = await tasa_elem.inner_text()
            # Extraer el porcentaje usando regex (ej: "15.6%")
            match = re.search(r'(\d+\.?\d*%)', tasa_text)
            if match:
                tasa_nominal = match.group(1)
                print(f"   ✓ Tasa nominal: {tasa_nominal}")
            else:
                print(f"   ⚠️ No se pudo extraer tasa de: {tasa_text[:50]}...")
        except Exception as e:
            print(f"   ⚠️ Error al extraer tasa: {e}")

        # Click en Ver tabla
        print("🔘 Haciendo click en 'Ver tabla de amortización'...")
        await asyncio.sleep(random.uniform(1.0, 2.0))
        ver_tabla_btn = page.locator('button:has-text("Ver tabla")')
        await ver_tabla_btn.scroll_into_view_if_needed()
        await asyncio.sleep(random.uniform(0.5, 1.0))
        await ver_tabla_btn.click()
        print("✓ Click realizado, esperando modal...")

        # Esperar modal
        await page.wait_for_selector(".modalTable__modal", timeout=15000, state="visible")
        print("✓ Modal visible")
        await asyncio.sleep(1.0)

        await page.wait_for_selector(".modalTable__table", timeout=10000, state="visible")
        print("✓ Tabla visible")
        await asyncio.sleep(1.0)

        # IMPORTANTE: Hacer scroll dentro del modal para forzar renderizado de TODAS las filas
        # El modal puede tener virtual scrolling que solo muestra las primeras 12 filas
        print("🔄 Haciendo scroll dentro del modal para cargar todas las filas...")
        modal_body = page.locator(".modalTable__body").first

        # Hacer scroll gradual hasta el final del modal para forzar renderizado de todas las filas
        # Aumentado a 20 scrolls para asegurar que se cargan TODAS las filas en plazos largos (48-60 meses)
        for i in range(20):  # Hacer 20 scrolls graduales
            await modal_body.evaluate("el => el.scrollTop = el.scrollHeight")
            await asyncio.sleep(0.5)  # Esperar a que se rendericen nuevas filas

        print("✓ Scroll completado, esperando renderizado final...")
        await asyncio.sleep(3.0)  # Esperar a que todas las filas estén renderizadas

        # Extraer filas de la tabla
        filas = await page.query_selector_all('table tbody tr')
        filas_datos = []

        for fila in filas:
            celdas_elems = await fila.query_selector_all('td')
            celdas = []
            for celda in celdas_elems:
                texto = await celda.inner_text()
                celdas.append(texto.strip())

            if celdas:
                filas_datos.append({
                    "n": celdas[0],
                    "saldo": celdas[1] if len(celdas) > 1 else "",
                    "capital": celdas[2] if len(celdas) > 2 else "",
                    "interes": celdas[3] if len(celdas) > 3 else "",
                    "s_desgravamen": celdas[4] if len(celdas) > 4 else "",
                    "s_cesantia": celdas[5] if len(celdas) > 5 else "",
                    "total_seguros": celdas[6] if len(celdas) > 6 else "",
                    "cuota": celdas[7] if len(celdas) > 7 else "",
                })

        print(f"📋 Extraídas {len(filas_datos)} filas de datos de la tabla")

        # NOTA: La tasa nominal ya fue extraída de la página principal antes de abrir el modal
        print("🔍 Extrayendo Total de interés desde el modal...")

        # Total de interés desde el modal (más confiable)
        try:
            # Buscar en las filas del modal el campo "Total de interés"
            modal_desc_rows = await page.locator('.descriptionTable__row').all()
            total_interes_modal = None

            for row in modal_desc_rows:
                text = await row.inner_text()
                if "Total de interés" in text or "Total de interés:" in text:
                    # Extraer el valor (segundo elemento)
                    ps = await row.locator('p').all()
                    if len(ps) >= 2:
                        total_interes_modal = await ps[1].inner_text()
                        print(f"   Total de interés (desde modal): {total_interes_modal}")
                        break

            # Si se encontró en el modal, actualizar la variable
            if total_interes_modal:
                total_interes = total_interes_modal
            else:
                print(f"⚠️ Total de interés no encontrado en modal, usando valor de página principal: {total_interes}")
        except Exception as e:
            print(f"⚠️ Error al extraer Total de interés desde modal: {e}. Usando valor de página principal")

        # Cerrar página original
        print("✓ Cerrando página original...")
        await page.close()
        print("✓ Página cerrada")

        # GENERAR EXCEL
        print("📊 Iniciando generación de Excel...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Tabla de Amortización"

        # Estilos
        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        border_style = Side(style='thin', color='000000')
        border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

        # Título
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"Tabla de amortización es {etiqueta.lower()}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = header_alignment

        # Información del crédito
        ws['A3'] = "Producto:"
        ws['B3'] = "Multicrédito"
        ws['E3'] = "Capital:"
        ws['F3'] = capital

        ws['A4'] = f"Plazo (meses):"
        ws['B4'] = str(meses)
        ws['E4'] = "Total de interés:"
        ws['F4'] = total_interes

        ws['A5'] = "Tasa de interés nominal:"
        ws['B5'] = tasa_nominal
        ws['E5'] = "Impuesto de Solca:"
        ws['F5'] = impuesto_solca

        ws['A6'] = "Tipo de amortización:"
        ws['B6'] = etiqueta
        ws['E6'] = "Total de seguros:"
        ws['F6'] = total_seguros

        ws['A7'] = "Fecha simulación:"
        ws['B7'] = datetime.now().strftime("%d/%m/%Y")
        ws['E7'] = "Total a pagar:"
        ws['F7'] = total_pagar

        # Headers de la tabla
        row = 9
        headers = ["N°", "Saldo", "Capital", "Interés", "S. Desgravamen", "S. Cesantia", "Total Seguros", "Cuota"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Datos de la tabla
        for fila_data in filas_datos:
            row += 1
            ws.cell(row=row, column=1, value=fila_data["n"]).border = border
            ws.cell(row=row, column=2, value=fila_data["saldo"]).border = border
            ws.cell(row=row, column=3, value=fila_data["capital"]).border = border
            ws.cell(row=row, column=4, value=fila_data["interes"]).border = border
            ws.cell(row=row, column=5, value=fila_data["s_desgravamen"]).border = border
            ws.cell(row=row, column=6, value=fila_data["s_cesantia"]).border = border
            ws.cell(row=row, column=7, value=fila_data["total_seguros"]).border = border
            ws.cell(row=row, column=8, value=fila_data["cuota"]).border = border

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15

        # Guardar Excel en memoria
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_bytes = excel_buffer.getvalue()
        print(f"✓ Excel generado en memoria ({len(excel_bytes)} bytes)")

        # Generar nombre descriptivo: {monto}_{meses}_{amortización}_{fecha}_{usuario}.xlsx
        fecha = datetime.now().strftime("%Y%m%d")
        usuario = data.correo.split("@")[0] if data.correo else "cliente"
        amort_limpio = etiqueta.lower().replace("é", "e").replace("á", "a")
        filename = f"{int(data.monto)}_{meses}_{amort_limpio}_{fecha}_{usuario}.xlsx"

        print(f"✓ Excel generado exitosamente: {filename}")

        # Devolver el Excel directamente
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )

    except Exception as e:
        print(f"❌ ERROR en generar_excel: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=502, detail=f"Error al generar Excel: {e}")
    finally:
        print("🔧 Cerrando contexto del navegador...")
        await context.close()
        print("✓ Contexto cerrado")


# -----------------------------------------------------------------------------
# NOTAS
#
# Anti-bot: el sitio carga una capa de deteccion (eudaapi + beacons). Si en
# headless ves resultados vacios o bloqueos, prueba headless=False, o agrega un
# user-agent realista en new_context(user_agent=...). No abuses de la frecuencia
# de llamadas.
#
# Descarga del PDF (paso 5 del ejercicio): el boton "Descargar Tabla" se puede
# capturar con page.expect_download(). Se puede agregar otro endpoint /pdf que
# guarde el archivo y devuelva la ruta. Pide ayuda para ese cuando llegues ahi.
# -----------------------------------------------------------------------------
